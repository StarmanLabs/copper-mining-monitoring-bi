"""Utilities to extract structured assumptions and benchmarks from the Excel model."""

from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet


@dataclass(frozen=True)
class WorkbookData:
    annual_inputs: pd.DataFrame
    assumptions: pd.DataFrame
    historical_prices: pd.DataFrame
    operational_history: pd.DataFrame
    benchmark_metrics: pd.DataFrame
    benchmark_distribution: pd.DataFrame


@dataclass(frozen=True)
class SheetSpec:
    key: str
    aliases: tuple[str, ...]


@dataclass(frozen=True)
class SeriesSpec:
    metric: str
    sheet_key: str
    year_col: int
    value_col: int
    row_start: int
    row_end: int
    unit: str


@dataclass(frozen=True)
class CellSpec:
    parameter: str
    sheet_key: str
    cell: str
    unit: str
    scale: float = 1.0
    note: str | None = None


@dataclass(frozen=True)
class BenchmarkSpec:
    metric: str
    sheet_key: str
    cell: str
    unit: str
    currency: str | None
    valuation_basis: str
    timing_basis: str
    note: str


class WorkbookStructureError(ValueError):
    """Raised when the workbook layout does not match the expected extraction map."""


SHEET_SPECS = (
    SheetSpec("market", ("Market_Data",)),
    SheetSpec("empirical", ("Empírical_Data", "Empirical_Data", "EmpÃ­rical_Data")),
    SheetSpec("expansion", ("Expansion_Model",)),
    SheetSpec("monte_carlo", ("Monte_carlo",)),
    SheetSpec("base_model", ("Base_Model",)),
    SheetSpec("sensitivity", ("Sensitivity",)),
)

ANNUAL_SERIES_SPECS = (
    SeriesSpec("copper_price_usd_per_lb", "market", 5, 6, 3, 17, "USD/lb"),
    SeriesSpec("base_processed_tonnes", "market", 12, 13, 3, 17, "tonnes"),
    SeriesSpec("base_head_grade", "market", 19, 20, 3, 17, "grade_decimal"),
    SeriesSpec("base_recovery", "market", 19, 20, 20, 34, "recovery_decimal"),
)

ASSUMPTION_SPECS = (
    CellSpec("mine_cost_usd_per_tonne", "market", "F22", "USD/t"),
    CellSpec("plant_cost_usd_per_tonne", "market", "F23", "USD/t"),
    CellSpec("g_and_a_cost_usd_per_tonne", "market", "F24", "USD/t"),
    CellSpec("base_unit_cost_usd_per_tonne", "market", "F25", "USD/t"),
    CellSpec("expansion_unit_cost_usd_per_tonne", "market", "G25", "USD/t"),
    CellSpec("income_tax_rate", "market", "M22", "decimal"),
    CellSpec(
        "royalty_rate",
        "market",
        "M23",
        "decimal",
        note="Listed in the workbook inputs, but not applied in the expansion cash-flow chain.",
    ),
    CellSpec(
        "special_levy_rate",
        "market",
        "M24",
        "decimal",
        note="Listed in the workbook inputs, but not applied in the expansion cash-flow chain.",
    ),
    CellSpec("wacc", "market", "M25", "decimal"),
    CellSpec("working_capital_ratio", "market", "M26", "decimal"),
    CellSpec("payable_rate", "market", "F29", "decimal"),
    CellSpec("tc_rc_usd_per_lb", "market", "F30", "USD/lb"),
    CellSpec("raw_initial_capex_year0_usd", "market", "M30", "USD", scale=1_000_000),
    CellSpec("raw_initial_capex_year1_usd", "market", "M31", "USD", scale=1_000_000),
    CellSpec(
        "raw_total_capex_reference_usd",
        "market",
        "M32",
        "USD",
        scale=1_000_000,
        note="Raw Market_Data capex reference. The active benchmark capex flows are driven by the Sensitivity sheet.",
    ),
    CellSpec("initial_capex_year0_usd", "sensitivity", "F3", "USD", scale=-1.0),
    CellSpec("initial_capex_year1_usd", "sensitivity", "F4", "USD", scale=-1.0),
    CellSpec(
        "sustaining_capex_usd",
        "expansion",
        "C84",
        "USD",
        scale=-1.0,
        note="Annual sustaining capex embedded in the expansion cash-flow block from year 2 onward.",
    ),
    CellSpec("expansion_uplift_pct", "expansion", "B1", "decimal"),
    CellSpec("expansion_midpoint_year", "expansion", "D1", "year"),
    CellSpec("expansion_steepness", "expansion", "E1", "decimal"),
    CellSpec("benchmark_npv_excel", "expansion", "L99", "USD", note="Workbook deterministic VAN for the incremental expansion flow."),
    CellSpec("benchmark_irr_excel", "expansion", "L101", "decimal", note="Workbook deterministic IRR for the incremental expansion flow."),
    CellSpec("mu_price_returns", "empirical", "P26", "decimal"),
    CellSpec("sigma_price_returns", "empirical", "P29", "decimal"),
    CellSpec("grade_trend_slope", "empirical", "I40", "grade_decimal"),
    CellSpec("grade_trend_intercept", "empirical", "J40", "grade_decimal"),
    CellSpec("average_recovery", "empirical", "L39", "decimal"),
    CellSpec("annual_decline_rate", "empirical", "O42", "decimal"),
)

BENCHMARK_SPECS = (
    BenchmarkSpec(
        metric="expected_npv",
        sheet_key="monte_carlo",
        cell="H17",
        unit="USD",
        currency="USD",
        valuation_basis="incremental_expansion",
        timing_basis="year0_initial_outlay_plus_discounted_year1_to_year15",
        note="Workbook Monte Carlo expected NPV for the incremental expansion model.",
    ),
    BenchmarkSpec(
        metric="median_npv",
        sheet_key="monte_carlo",
        cell="H20",
        unit="USD",
        currency="USD",
        valuation_basis="incremental_expansion",
        timing_basis="year0_initial_outlay_plus_discounted_year1_to_year15",
        note="Workbook Monte Carlo median NPV for the incremental expansion model.",
    ),
    BenchmarkSpec(
        metric="std_npv",
        sheet_key="monte_carlo",
        cell="H23",
        unit="USD",
        currency="USD",
        valuation_basis="incremental_expansion",
        timing_basis="year0_initial_outlay_plus_discounted_year1_to_year15",
        note="Workbook Monte Carlo standard deviation for the incremental expansion model.",
    ),
    BenchmarkSpec(
        metric="min_npv",
        sheet_key="monte_carlo",
        cell="H26",
        unit="USD",
        currency="USD",
        valuation_basis="incremental_expansion",
        timing_basis="year0_initial_outlay_plus_discounted_year1_to_year15",
        note="Workbook Monte Carlo minimum NPV for the incremental expansion model.",
    ),
    BenchmarkSpec(
        metric="max_npv",
        sheet_key="monte_carlo",
        cell="H29",
        unit="USD",
        currency="USD",
        valuation_basis="incremental_expansion",
        timing_basis="year0_initial_outlay_plus_discounted_year1_to_year15",
        note="Workbook Monte Carlo maximum NPV for the incremental expansion model.",
    ),
    BenchmarkSpec(
        metric="cv_npv",
        sheet_key="monte_carlo",
        cell="H32",
        unit="ratio",
        currency=None,
        valuation_basis="incremental_expansion",
        timing_basis="year0_initial_outlay_plus_discounted_year1_to_year15",
        note="Workbook Monte Carlo coefficient of variation for the incremental expansion model.",
    ),
    BenchmarkSpec(
        metric="var_5pct",
        sheet_key="monte_carlo",
        cell="H35",
        unit="USD",
        currency="USD",
        valuation_basis="incremental_expansion",
        timing_basis="year0_initial_outlay_plus_discounted_year1_to_year15",
        note="Workbook 5th percentile NPV for the incremental expansion model.",
    ),
    BenchmarkSpec(
        metric="cvar_5pct",
        sheet_key="monte_carlo",
        cell="H38",
        unit="USD",
        currency="USD",
        valuation_basis="incremental_expansion",
        timing_basis="year0_initial_outlay_plus_discounted_year1_to_year15",
        note="Workbook conditional tail mean beyond the 5th percentile for the incremental expansion model.",
    ),
)


def _resolve_sheet(workbook: Workbook, spec: SheetSpec) -> Worksheet:
    available = {name.casefold(): name for name in workbook.sheetnames}
    for alias in spec.aliases:
        if alias.casefold() in available:
            return workbook[available[alias.casefold()]]
    expected = ", ".join(spec.aliases)
    found = ", ".join(workbook.sheetnames)
    raise WorkbookStructureError(
        f"Required worksheet for '{spec.key}' not found. Expected one of [{expected}] but workbook contains [{found}]."
    )


def _read_required_cell(worksheet: Worksheet, cell: str) -> float:
    value = worksheet[cell].value
    if value is None:
        raise WorkbookStructureError(
            f"Required cell {worksheet.title}!{cell} is empty. Review the workbook structure or update the extraction map."
        )
    return float(value)


def _series_from_rows(worksheet: Worksheet, spec: SeriesSpec) -> pd.DataFrame:
    records: list[dict[str, float | int | str]] = []
    for row in range(spec.row_start, spec.row_end + 1):
        year = worksheet.cell(row, spec.year_col).value
        value = worksheet.cell(row, spec.value_col).value
        if year is None or value is None:
            raise WorkbookStructureError(
                f"Required series cell missing while loading '{spec.metric}' at {worksheet.title}!R{row}C{spec.value_col}."
            )
        records.append(
            {
                "year": int(year),
                "metric": spec.metric,
                "value": float(value),
                "unit": spec.unit,
                "source_sheet": worksheet.title,
            }
        )
    return pd.DataFrame(records)


def _build_assumptions(worksheets: dict[str, Worksheet]) -> pd.DataFrame:
    records = []
    for spec in ASSUMPTION_SPECS:
        worksheet = worksheets[spec.sheet_key]
        records.append(
            {
                "parameter": spec.parameter,
                "value": _read_required_cell(worksheet, spec.cell) * spec.scale,
                "unit": spec.unit,
                "source_sheet": worksheet.title,
                "source_cell": spec.cell,
                "note": spec.note,
            }
        )
    return pd.DataFrame(records)


def _build_benchmark_metrics(worksheets: dict[str, Worksheet]) -> pd.DataFrame:
    records = []
    for spec in BENCHMARK_SPECS:
        worksheet = worksheets[spec.sheet_key]
        records.append(
            {
                "metric": spec.metric,
                "value": _read_required_cell(worksheet, spec.cell),
                "unit": spec.unit,
                "currency": spec.currency,
                "valuation_basis": spec.valuation_basis,
                "timing_basis": spec.timing_basis,
                "source_sheet": worksheet.title,
                "source_cell": spec.cell,
                "note": spec.note,
            }
        )
    return pd.DataFrame(records)


def _build_historical_prices(empirical: Worksheet) -> pd.DataFrame:
    historical_price_records = []
    for row in range(3, 24):
        year = empirical.cell(row, 1).value
        usd_per_mt = empirical.cell(row, 2).value
        usd_per_lb = empirical.cell(row, 4).value
        log_return = empirical.cell(row, 16).value
        if year is None:
            continue
        historical_price_records.append(
            {
                "calendar_year": int(year),
                "copper_price_usd_per_mt": float(usd_per_mt),
                "copper_price_usd_per_lb": float(usd_per_lb),
                "log_return": None if log_return is None else float(log_return),
            }
        )
    return pd.DataFrame(historical_price_records)


def _build_operational_history(empirical: Worksheet) -> pd.DataFrame:
    operational_records = []
    for row in range(44, 55):
        calendar_year = empirical.cell(row, 1).value
        grade_decimal = empirical.cell(row, 3).value
        recovery_decimal = empirical.cell(row, 5).value
        if calendar_year is None:
            continue
        operational_records.append(
            {
                "calendar_year": int(calendar_year),
                "head_grade": float(grade_decimal),
                "recovery": float(recovery_decimal),
            }
        )
    return pd.DataFrame(operational_records)


def _build_benchmark_distribution(monte_carlo: Worksheet) -> pd.DataFrame:
    distribution_records = []
    for row in range(17, 10017):
        npv_sorted = monte_carlo.cell(row, 11).value
        cumulative_probability = monte_carlo.cell(row, 12).value
        if npv_sorted is None or cumulative_probability is None:
            continue
        distribution_records.append(
            {
                "npv_sorted": float(npv_sorted),
                "cumulative_probability": float(cumulative_probability),
                "unit": "USD",
                "currency": "USD",
                "valuation_basis": "incremental_expansion",
            }
        )
    return pd.DataFrame(distribution_records)


def _validate_annual_inputs(annual_inputs: pd.DataFrame) -> None:
    years = annual_inputs["year"].tolist()
    expected_years = list(range(1, len(years) + 1))
    if years != expected_years:
        raise WorkbookStructureError(
            f"Annual project years must be consecutive integers starting at 1. Found {years}."
        )
    if (annual_inputs["copper_price_usd_per_lb"] <= 0).any():
        raise WorkbookStructureError("Copper price series contains non-positive values.")
    if (annual_inputs["base_processed_tonnes"] <= 0).any():
        raise WorkbookStructureError("Processed tonnes series contains non-positive values.")
    if (~annual_inputs["base_head_grade"].between(0.0, 1.0)).any():
        raise WorkbookStructureError("Head grade series contains values outside the [0, 1] range.")
    if (~annual_inputs["base_recovery"].between(0.0, 1.0)).any():
        raise WorkbookStructureError("Recovery series contains values outside the [0, 1] range.")


def _validate_assumptions(assumptions: pd.DataFrame) -> None:
    values = assumptions.set_index("parameter")["value"]
    bounded_parameters = {
        "payable_rate": (0.0, 1.0),
        "income_tax_rate": (0.0, 1.0),
        "royalty_rate": (0.0, 1.0),
        "special_levy_rate": (0.0, 1.0),
        "wacc": (0.0, 1.0),
        "working_capital_ratio": (0.0, 1.0),
    }
    for parameter, (lower, upper) in bounded_parameters.items():
        value = float(values[parameter])
        if not lower <= value <= upper:
            raise WorkbookStructureError(
                f"Assumption '{parameter}' is outside the expected range [{lower}, {upper}]. Found {value}."
            )

    non_negative_parameters = (
        "mine_cost_usd_per_tonne",
        "plant_cost_usd_per_tonne",
        "g_and_a_cost_usd_per_tonne",
        "base_unit_cost_usd_per_tonne",
        "expansion_unit_cost_usd_per_tonne",
        "initial_capex_year0_usd",
        "initial_capex_year1_usd",
        "sustaining_capex_usd",
        "tc_rc_usd_per_lb",
    )
    for parameter in non_negative_parameters:
        value = float(values[parameter])
        if value < 0:
            raise WorkbookStructureError(f"Assumption '{parameter}' must be non-negative. Found {value}.")


def _validate_benchmark_distribution(benchmark_distribution: pd.DataFrame) -> None:
    if benchmark_distribution.empty:
        raise WorkbookStructureError("Benchmark distribution table is empty.")
    cumulative_probability = benchmark_distribution["cumulative_probability"]
    if (~cumulative_probability.between(0.0, 1.0)).any():
        raise WorkbookStructureError("Benchmark cumulative probabilities must stay within the [0, 1] range.")
    if not cumulative_probability.is_monotonic_increasing:
        raise WorkbookStructureError("Benchmark cumulative probabilities must be monotonically increasing.")


def load_workbook_data(workbook_path: str | Path) -> WorkbookData:
    workbook_path = Path(workbook_path)
    wb = load_workbook(workbook_path, data_only=True, keep_vba=True)
    worksheets = {spec.key: _resolve_sheet(wb, spec) for spec in SHEET_SPECS}

    assumptions = _build_assumptions(worksheets)
    assumptions_series = assumptions.set_index("parameter")["value"]

    annual_series = {
        spec.metric: _series_from_rows(worksheets[spec.sheet_key], spec) for spec in ANNUAL_SERIES_SPECS
    }
    annual_inputs = (
        annual_series["copper_price_usd_per_lb"][["year", "value"]]
        .rename(columns={"value": "copper_price_usd_per_lb"})
        .merge(
            annual_series["base_processed_tonnes"][["year", "value"]].rename(columns={"value": "base_processed_tonnes"}),
            on="year",
            how="inner",
            validate="one_to_one",
        )
        .merge(
            annual_series["base_head_grade"][["year", "value"]].rename(columns={"value": "base_head_grade"}),
            on="year",
            how="inner",
            validate="one_to_one",
        )
        .merge(
            annual_series["base_recovery"][["year", "value"]].rename(columns={"value": "base_recovery"}),
            on="year",
            how="inner",
            validate="one_to_one",
        )
        .sort_values("year")
        .reset_index(drop=True)
    )
    annual_inputs["net_price_usd_per_lb"] = (
        annual_inputs["copper_price_usd_per_lb"] * assumptions_series["payable_rate"]
        - assumptions_series["tc_rc_usd_per_lb"]
    )
    _validate_assumptions(assumptions)
    _validate_annual_inputs(annual_inputs)

    historical_prices = _build_historical_prices(worksheets["empirical"])
    operational_history = _build_operational_history(worksheets["empirical"])
    benchmark_metrics = _build_benchmark_metrics(worksheets)
    benchmark_distribution = _build_benchmark_distribution(worksheets["monte_carlo"])
    _validate_benchmark_distribution(benchmark_distribution)

    return WorkbookData(
        annual_inputs=annual_inputs,
        assumptions=assumptions,
        historical_prices=historical_prices,
        operational_history=operational_history,
        benchmark_metrics=benchmark_metrics,
        benchmark_distribution=benchmark_distribution,
    )
